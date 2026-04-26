"""File system tools: read and write files."""
from __future__ import annotations

import json
import re
import shutil
from pathlib import Path
from urllib.parse import urlparse

from hushclaw.tools.base import tool, ToolResult

_ARTIFACTS_DIRNAME = "artifacts"


def _build_absolute_url(rel_url: str, _config=None) -> str:
    """Return an absolute public URL when configured."""
    base_url = ""
    if _config is not None:
        base_url = str(getattr(_config.server, "public_base_url", "") or "").strip()
    if not base_url:
        return ""
    parsed = urlparse(base_url)
    if parsed.scheme not in ("http", "https") or not parsed.netloc:
        return ""
    return f"{base_url.rstrip('/')}{rel_url}"


def _normalize_relative_path(path: str | Path, *, field_name: str) -> str:
    """Normalize a relative artifact path and reject traversal."""
    rel = Path(path)
    if rel.is_absolute() or ".." in rel.parts:
        raise ValueError(f"Invalid {field_name}: {path}")
    rel_posix = rel.as_posix().strip("/")
    if not rel_posix:
        raise ValueError(f"Invalid {field_name}: {path}")
    return rel_posix


def _build_artifact_meta(
    artifact_id: str,
    *,
    kind: str,
    name: str,
    relative_path: str,
    _config=None,
) -> dict:
    """Build normalized artifact metadata for web UI consumption."""
    rel_path = _normalize_relative_path(relative_path, field_name="artifact path")
    root_url = f"/files/{_ARTIFACTS_DIRNAME}/{artifact_id}/"
    rel_url = f"{root_url}{rel_path}"
    meta = {
        "trusted": True,
        "kind": kind,
        "url": rel_url,
        "name": name,
        "artifact_id": artifact_id,
        "file_id": artifact_id,
        "root_url": root_url,
    }
    if kind == "directory":
        meta["entry_url"] = rel_url
        meta["entry_name"] = Path(rel_path).name
    absolute_url = _build_absolute_url(rel_url, _config=_config)
    if absolute_url:
        meta["absolute_url"] = absolute_url
        absolute_root = _build_absolute_url(root_url, _config=_config)
        if absolute_root:
            meta["absolute_root_url"] = absolute_root
        if kind == "directory":
            meta["absolute_entry_url"] = absolute_url
    return meta


def _get_upload_dir(_config) -> Path:
    upload_dir: Path | None = None
    if _config is not None:
        upload_dir = getattr(_config.server, "upload_dir", None)
    if upload_dir is None:
        raise ValueError("upload_dir not configured — cannot generate artifact URL")
    upload_dir = Path(upload_dir)
    upload_dir.mkdir(parents=True, exist_ok=True)
    return upload_dir


def _get_artifacts_dir(_config) -> Path:
    artifacts_dir = _get_upload_dir(_config) / _ARTIFACTS_DIRNAME
    artifacts_dir.mkdir(parents=True, exist_ok=True)
    return artifacts_dir


def _reuse_directory_artifact(src_dir: Path, artifacts_dir: Path, entrypoint: str) -> dict | None:
    """Return existing artifact coordinates when src_dir already lives under artifacts."""
    try:
        rel = src_dir.resolve().relative_to(artifacts_dir.resolve())
    except Exception:
        return None
    if len(rel.parts) != 1:
        return None
    artifact_id = rel.parts[0]
    entry_rel = _normalize_relative_path(entrypoint, field_name="artifact entrypoint")
    return {
        "artifact_id": artifact_id,
        "relative_path": entry_rel,
    }


def _reuse_file_artifact(src: Path, artifacts_dir: Path) -> dict | None:
    """Return existing artifact coordinates when src already lives under artifacts."""
    try:
        rel = src.resolve().relative_to(artifacts_dir.resolve())
    except Exception:
        return None
    if len(rel.parts) < 2:
        return None
    artifact_id = rel.parts[0]
    rel_path = Path(*rel.parts[1:]).as_posix()
    return {
        "artifact_id": artifact_id,
        "relative_path": rel_path,
    }


def register_download_bundle(path: str | Path, _config=None, entrypoint: str = "index.html",
                             display_name: str = "") -> dict:
    """Copy a local directory into upload_dir/artifacts and return an entry URL."""
    from uuid import uuid4

    src_dir = Path(path).expanduser()
    if not src_dir.exists():
        raise FileNotFoundError(f"Path not found: {src_dir}")
    if not src_dir.is_dir():
        raise NotADirectoryError(f"Not a directory: {src_dir}")

    entry_rel = _normalize_relative_path(entrypoint, field_name="artifact entrypoint")

    src_entry = src_dir / entry_rel
    if not src_entry.exists() or not src_entry.is_file():
        raise FileNotFoundError(f"Artifact entrypoint not found: {src_entry}")

    artifacts_dir = _get_artifacts_dir(_config)

    reused = _reuse_directory_artifact(src_dir, artifacts_dir, entry_rel)
    if reused is not None:
        return _build_artifact_meta(
            reused["artifact_id"],
            kind="directory",
            name=display_name or src_dir.name,
            relative_path=reused["relative_path"],
            _config=_config,
        )

    artifact_id = uuid4().hex[:12]
    dest_root = artifacts_dir / artifact_id
    shutil.copytree(src_dir, dest_root)
    return _build_artifact_meta(
        artifact_id,
        kind="directory",
        name=display_name or src_dir.name,
        relative_path=entry_rel,
        _config=_config,
    )


def register_download_path(path: str | Path, _config=None, display_name: str = "") -> dict:
    """Register a file or directory artifact and return normalized metadata."""
    from uuid import uuid4

    src = Path(path).expanduser()
    if not src.exists():
        raise FileNotFoundError(f"Path not found: {src}")
    if src.is_dir():
        # Pick the best entrypoint: prefer index.html, fall back to first file.
        entrypoint = "index.html"
        if not (src / entrypoint).is_file():
            first = next((f for f in sorted(src.rglob("*")) if f.is_file()), None)
            if first is None:
                raise FileNotFoundError(f"Directory is empty, cannot register as artifact: {src}")
            entrypoint = first.relative_to(src).as_posix()
        return register_download_bundle(src, _config=_config, display_name=display_name, entrypoint=entrypoint)
    if not src.is_file():
        raise IsADirectoryError(f"Not a file: {src}")

    artifacts_dir = _get_artifacts_dir(_config)

    reused = _reuse_file_artifact(src, artifacts_dir)
    if reused is not None:
        return _build_artifact_meta(
            reused["artifact_id"],
            kind="file",
            name=display_name or src.name,
            relative_path=reused["relative_path"],
            _config=_config,
        )

    safe_name = re.sub(r"[^\w.\-]", "_", display_name or src.name)[:128] or "file"
    artifact_id = uuid4().hex[:12]
    dest_root = artifacts_dir / artifact_id
    dest_root.mkdir(parents=True, exist_ok=True)
    dest = dest_root / safe_name
    shutil.copy2(src, dest)
    return _build_artifact_meta(
        artifact_id,
        kind="file",
        name=safe_name,
        relative_path=safe_name,
        _config=_config,
    )


def _read_pdf(p: Path, max_chars: int) -> tuple[str, bool]:
    import pdfplumber  # type: ignore
    parts = []
    with pdfplumber.open(str(p)) as pdf:
        for pg in pdf.pages:
            t = pg.extract_text()
            if t:
                parts.append(t)
    text = "\n\n".join(parts)
    return text[:max_chars], len(text) > max_chars


def _read_word(p: Path, max_chars: int) -> tuple[str, bool]:
    from docx import Document  # type: ignore
    doc = Document(str(p))
    text = "\n".join(para.text for para in doc.paragraphs if para.text.strip())
    return text[:max_chars], len(text) > max_chars


def _read_excel(p: Path, max_chars: int) -> tuple[str, bool]:
    from openpyxl import load_workbook  # type: ignore
    wb = load_workbook(str(p), read_only=True, data_only=True)
    parts = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            row_str = "\t".join(str(c) if c is not None else "" for c in row)
            if row_str.strip():
                rows.append(row_str)
        if rows:
            parts.append(f"[Sheet: {name}]\n" + "\n".join(rows))
    text = "\n\n".join(parts)
    return text[:max_chars], len(text) > max_chars


@tool(
    name="read_file",
    description=(
        "Read the contents of a file at the specified path. "
        "Supports plain text, Markdown, CSV, PDF, Word (.docx), and Excel (.xlsx). "
        "Binary formats are extracted to readable text automatically."
    ),
    parallel_safe=True,
)
def read_file(path: str, max_chars: int = 32768) -> ToolResult:
    """Read a file; auto-extracts text from PDF/Word/Excel."""
    try:
        p = Path(path).expanduser()
        if not p.exists():
            return ToolResult.error(f"File not found: {path}")
        if not p.is_file():
            return ToolResult.error(f"Not a file: {path}")

        ext = p.suffix.lower()

        if ext == ".pdf":
            try:
                text, truncated = _read_pdf(p, max_chars)
            except ImportError:
                return ToolResult.error("pdfplumber is not installed. Run: pip install pdfplumber")
            prefix = f"[Truncated to {max_chars} chars]\n" if truncated else ""
            return ToolResult.ok(prefix + text)

        if ext in (".docx", ".doc"):
            try:
                text, truncated = _read_word(p, max_chars)
            except ImportError:
                return ToolResult.error("python-docx is not installed. Run: pip install python-docx")
            prefix = f"[Truncated to {max_chars} chars]\n" if truncated else ""
            return ToolResult.ok(prefix + text)

        if ext in (".xlsx", ".xls", ".xlsm"):
            try:
                text, truncated = _read_excel(p, max_chars)
            except ImportError:
                return ToolResult.error("openpyxl is not installed. Run: pip install openpyxl")
            prefix = f"[Truncated to {max_chars} chars]\n" if truncated else ""
            return ToolResult.ok(prefix + text)

        # Plain text fallback (txt, md, csv, json, code, etc.)
        size = p.stat().st_size
        if size > max_chars:
            content = p.read_bytes()[:max_chars].decode("utf-8", errors="replace")
            return ToolResult.ok(f"[Truncated to {max_chars} chars]\n{content}")
        return ToolResult.ok(p.read_text(encoding="utf-8", errors="replace"))

    except PermissionError:
        return ToolResult.error(f"Permission denied: {path}")
    except Exception as e:
        return ToolResult.error(f"Failed to read {path}: {e}")


@tool(
    name="write_file",
    description=(
        "Write content to a file. Use a relative path (e.g. 'report.md') to write inside "
        "the active workspace's files directory — this is the preferred location for generated "
        "files. Absolute paths (~/... or /...) are also accepted for other locations. "
        "Do NOT use /files/ as a path — that is a URL prefix, not a filesystem directory. "
        "Returns the file path and a /files/ download URL so the user can access the file."
    ),
)
def write_file(path: str, content: str, _config=None) -> ToolResult:
    """Write content to a file and return a download URL."""
    try:
        p = Path(path).expanduser()
        if not p.is_absolute():
            ws_dir = getattr(getattr(_config, "agent", None), "workspace_dir", None) if _config else None
            base = Path(ws_dir) / "files" if ws_dir else Path.home() / "Downloads"
            p = base / path
        if path.startswith("/files/"):
            return ToolResult.error(
                "'/files/' paths are read-only served URLs. "
                "Write to a normal filesystem path, then register it with make_download_url."
            )

        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(content, encoding="utf-8")

        # Auto-register as downloadable artifact when server config is available.
        if _config is not None:
            try:
                meta = register_download_path(p, _config=_config)
                url = meta.get("absolute_url") or meta.get("url", "")
                result = ToolResult.ok(f"Written {len(content)} chars to {p}\nDownload: {url}")
                result.artifact_id = meta.get("artifact_id", "")
                return result
            except Exception:
                pass  # Fall through to plain success message if registration fails

        return ToolResult.ok(f"Written {len(content)} characters to {p}")
    except PermissionError:
        return ToolResult.error(f"Permission denied: {path}")
    except Exception as e:
        return ToolResult.error(f"Failed to write {path}: {e}")


@tool(
    name="list_dir",
    description="List files and directories at a given path.",
    parallel_safe=True,
)
def list_dir(path: str = ".") -> ToolResult:
    """List directory contents."""
    try:
        p = Path(path).expanduser()
        if not p.exists():
            return ToolResult.error(f"Path not found: {path}")
        if not p.is_dir():
            return ToolResult.error(f"Not a directory: {path}")
        entries = sorted(p.iterdir(), key=lambda e: (e.is_file(), e.name))
        lines = []
        for entry in entries:
            kind = "F" if entry.is_file() else "D"
            lines.append(f"[{kind}] {entry.name}")
        return ToolResult.ok("\n".join(lines) if lines else "(empty)")
    except Exception as e:
        return ToolResult.error(f"Failed to list {path}: {e}")


@tool(
    name="make_download_url",
    description=(
        "Register a local file or directory as a downloadable / previewable artifact "
        "and return its /files/ URL."
    ),
    parallel_safe=True,
)
def make_download_url(path: str, _config=None) -> ToolResult:
    """Register a file or directory artifact and return a /files/ URL."""
    try:
        payload = register_download_path(path, _config=_config)
        return ToolResult.ok(json.dumps(payload, ensure_ascii=False))
    except Exception as e:
        return ToolResult.error(f"Failed to register artifact: {e}")


@tool(
    name="make_download_bundle",
    description=(
        "Register a local directory as a previewable artifact and return the "
        "entry-page URL. Use this when generated output depends on sibling assets."
    ),
    parallel_safe=True,
)
def make_download_bundle(path: str, entrypoint: str = "index.html", _config=None) -> ToolResult:
    """Copy a directory into upload_dir/artifacts and return its entry-page URL."""
    try:
        payload = register_download_bundle(path, _config=_config, entrypoint=entrypoint)
        return ToolResult.ok(json.dumps(payload, ensure_ascii=False))
    except Exception as e:
        return ToolResult.error(f"Failed to register artifact directory: {e}")
